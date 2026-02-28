from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── colour palette ──────────────────────────────────────────────────────────
C_DARK    = "1A1A2E"   # very dark navy – headers
C_MID     = "16213E"   # dark navy – sub-headers
C_ACCENT  = "0F3460"   # medium navy
C_GOLD    = "E94560"   # red-gold accent
C_LIGHT   = "F5F5F5"   # near-white rows
C_WHITE   = "FFFFFF"
C_BLUE_IN = "D6E4F0"   # light blue input cells
C_WARN    = "FFF3CD"   # warning yellow for core/structure
C_GREEN   = "D4EDDA"   # pass green
C_RED     = "F8D7DA"   # fail red

def hdr_fill(hex_col):
    return PatternFill("solid", start_color=hex_col, fgColor=hex_col)

def side(style="thin", color="BBBBBB"):
    return Side(style=style, color=color)

thin_border = Border(left=side(), right=side(), top=side(), bottom=side())
thick_bottom = Border(bottom=Side(style="medium", color="333333"))

def h_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def b_font(size=10, bold=False, color="222222"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right_align():
    return Alignment(horizontal="right", vertical="center")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – ZONE AREA MATRIX
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Zone Area Matrix"

zones = [
    ("MEDICAL",           35000),
    ("HOTEL",             50000),
    ("TRANSPORTATION HQ", 52000),
    ("CORPORATE OFFICE",  30000),
    ("ADMIN OFFICE",      30000),
    ("ENTERTAINMENT",     31000),
    ("SKY ZONE",          10400),
]

# ── title block ─────────────────────────────────────────────────────────────
ws1.merge_cells("A1:I1")
ws1["A1"] = "ZONE AREA DISTRIBUTION MATRIX"
ws1["A1"].font = Font(name="Arial", size=16, bold=True, color=C_WHITE)
ws1["A1"].fill = hdr_fill(C_DARK)
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:I2")
ws1["A2"] = "Core Deduction: 15%  |  Structural Deduction: 15%  |  Net Usable = 70% of Gross"
ws1["A2"].font = Font(name="Arial", size=10, italic=True, color=C_WHITE)
ws1["A2"].fill = hdr_fill(C_ACCENT)
ws1["A2"].alignment = center()
ws1.row_dimensions[2].height = 20

# ── assumptions block ────────────────────────────────────────────────────────
ws1["A4"] = "ASSUMPTIONS"
ws1["A4"].font = h_font(10, True, C_WHITE)
ws1["A4"].fill = hdr_fill(C_MID)
ws1["A4"].alignment = left()

labels_vals = [
    ("Core %",       "A5", "B5", 0.15),
    ("Structural %", "A6", "B6", 0.15),
]
for lbl, la, lb, val in labels_vals:
    ws1[la] = lbl
    ws1[la].font = b_font(10, True)
    ws1[la].fill = PatternFill("solid", start_color=C_WARN, fgColor=C_WARN)
    ws1[la].alignment = left()
    ws1[lb] = val
    ws1[lb].number_format = "0%"
    ws1[lb].font = Font(name="Arial", size=10, bold=True, color="0000CC")
    ws1[lb].fill = PatternFill("solid", start_color=C_BLUE_IN, fgColor=C_BLUE_IN)
    ws1[lb].alignment = center()

ws1["A7"] = "Net Usable %"
ws1["A7"].font = b_font(10, True)
ws1["B7"] = "=1-B5-B6"
ws1["B7"].number_format = "0%"
ws1["B7"].font = Font(name="Arial", size=10, bold=True, color="006600")
ws1["B7"].alignment = center()

# ── column headers ───────────────────────────────────────────────────────────
row_h = 9
headers = [
    "ZONE", "GROSS AREA\n(m²)", "CORE AREA\n(m²)",
    "STRUCTURAL\nAREA (m²)", "TOTAL\nDEDUCTION (m²)",
    "NET USABLE\nAREA (m²)", "CORE %", "STRUCTURAL %", "NET %"
]
col_widths = [22, 16, 16, 16, 18, 18, 10, 13, 10]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=row_h, column=ci, value=h)
    cell.font = h_font(10)
    cell.fill = hdr_fill(C_MID)
    cell.alignment = center()
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(ci)].width = w

ws1.row_dimensions[row_h].height = 40

# ── data rows ────────────────────────────────────────────────────────────────
zone_rows = {}   # zone_name -> excel row number
for i, (zone, gross) in enumerate(zones):
    r = row_h + 1 + i
    fill_col = C_LIGHT if i % 2 == 0 else C_WHITE

    # A – zone name
    ws1.cell(r, 1, zone).font = Font(name="Arial", size=10, bold=True)
    ws1.cell(r, 1).fill = PatternFill("solid", start_color=C_DARK, fgColor=C_DARK)
    ws1.cell(r, 1).font = Font(name="Arial", size=10, bold=True, color=C_GOLD)
    ws1.cell(r, 1).alignment = left()
    ws1.cell(r, 1).border = thin_border

    # B – gross area (input, blue)
    ws1.cell(r, 2, gross).font = Font(name="Arial", size=10, bold=True, color="0000CC")
    ws1.cell(r, 2).fill = PatternFill("solid", start_color=C_BLUE_IN, fgColor=C_BLUE_IN)
    ws1.cell(r, 2).number_format = "#,##0"
    ws1.cell(r, 2).alignment = center()
    ws1.cell(r, 2).border = thin_border

    b_ref = f"B{r}"
    # C – core
    ws1.cell(r, 3, f"={b_ref}*$B$5").font = b_font()
    ws1.cell(r, 3).fill = PatternFill("solid", start_color=C_WARN, fgColor=C_WARN)
    ws1.cell(r, 3).number_format = "#,##0"
    ws1.cell(r, 3).alignment = center()
    ws1.cell(r, 3).border = thin_border

    # D – structural
    ws1.cell(r, 4, f"={b_ref}*$B$6").font = b_font()
    ws1.cell(r, 4).fill = PatternFill("solid", start_color=C_WARN, fgColor=C_WARN)
    ws1.cell(r, 4).number_format = "#,##0"
    ws1.cell(r, 4).alignment = center()
    ws1.cell(r, 4).border = thin_border

    # E – total deduction
    ws1.cell(r, 5, f"=C{r}+D{r}").font = b_font(bold=True)
    ws1.cell(r, 5).fill = PatternFill("solid", start_color="FFD9D9", fgColor="FFD9D9")
    ws1.cell(r, 5).number_format = "#,##0"
    ws1.cell(r, 5).alignment = center()
    ws1.cell(r, 5).border = thin_border

    # F – net usable
    ws1.cell(r, 6, f"={b_ref}*$B$7").font = Font(name="Arial", size=10, bold=True, color="006600")
    ws1.cell(r, 6).fill = PatternFill("solid", start_color=C_GREEN, fgColor=C_GREEN)
    ws1.cell(r, 6).number_format = "#,##0"
    ws1.cell(r, 6).alignment = center()
    ws1.cell(r, 6).border = thin_border

    # G/H/I – percentages (formula-driven)
    ws1.cell(r, 7, "=$B$5").number_format = "0%"
    ws1.cell(r, 8, "=$B$6").number_format = "0%"
    ws1.cell(r, 9, "=$B$7").number_format = "0%"
    for ci in (7, 8, 9):
        ws1.cell(r, ci).font = b_font()
        ws1.cell(r, ci).fill = PatternFill("solid", start_color=fill_col, fgColor=fill_col)
        ws1.cell(r, ci).alignment = center()
        ws1.cell(r, ci).border = thin_border

    ws1.row_dimensions[r].height = 22
    zone_rows[zone] = r

# ── totals row ───────────────────────────────────────────────────────────────
r_tot = row_h + 1 + len(zones)
ws1.cell(r_tot, 1, "TOTAL").font = h_font(11)
ws1.cell(r_tot, 1).fill = hdr_fill(C_GOLD)
ws1.cell(r_tot, 1).alignment = center()
ws1.cell(r_tot, 1).border = thin_border

for ci in range(2, 7):
    col = get_column_letter(ci)
    first_r = row_h + 1
    last_r  = r_tot - 1
    ws1.cell(r_tot, ci, f"=SUM({col}{first_r}:{col}{last_r})")
    ws1.cell(r_tot, ci).font = h_font(11)
    ws1.cell(r_tot, ci).fill = hdr_fill(C_DARK)
    ws1.cell(r_tot, ci).number_format = "#,##0"
    ws1.cell(r_tot, ci).alignment = center()
    ws1.cell(r_tot, ci).border = thin_border

for ci in (7, 8, 9):
    ws1.cell(r_tot, ci, "").fill = hdr_fill(C_DARK)
    ws1.cell(r_tot, ci).border = thin_border

ws1.row_dimensions[r_tot].height = 26

# freeze pane
ws1.freeze_panes = "B10"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – PROGRAM THRESHOLD MATRIX  (JSON-ready for the automate function)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Threshold Matrix")

ws2.merge_cells("A1:E1")
ws2["A1"] = "PROGRAM THRESHOLD MATRIX  —  Input for Speckle Automate"
ws2["A1"].font = h_font(13)
ws2["A1"].fill = hdr_fill(C_DARK)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A2:E2")
ws2["A2"] = (
    "Max % = maximum share one program may occupy before the floor is flagged as mono-functional"
)
ws2["A2"].font = Font(name="Arial", size=9, italic=True, color=C_WHITE)
ws2["A2"].fill = hdr_fill(C_ACCENT)
ws2["A2"].alignment = center()

thresh_headers = ["ZONE / PROGRAM", "GROSS AREA (m²)", "NET USABLE (m²)", "MAX % (Threshold)", "NOTES"]
thresh_widths  = [24, 18, 18, 18, 30]
for ci, (h, w) in enumerate(zip(thresh_headers, thresh_widths), 1):
    ws2.cell(4, ci, h).font = h_font(10)
    ws2.cell(4, ci).fill = hdr_fill(C_MID)
    ws2.cell(4, ci).alignment = center()
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[4].height = 30

# default thresholds (tunable)
thresholds = {
    "MEDICAL":           ("'Zone Area Matrix'!B10", "'Zone Area Matrix'!F10", 75, "Clinical + admin mix; no single use >75%"),
    "HOTEL":             ("'Zone Area Matrix'!B11", "'Zone Area Matrix'!F11", 80, "Hotel keys dominant; F&B/support allowed up to 20%"),
    "TRANSPORTATION HQ": ("'Zone Area Matrix'!B12", "'Zone Area Matrix'!F12", 70, "Ops + office mix; mono-use flagged >70%"),
    "CORPORATE OFFICE":  ("'Zone Area Matrix'!B13", "'Zone Area Matrix'!F13", 75, "Open-plan office; support uses <25%"),
    "ADMIN OFFICE":      ("'Zone Area Matrix'!B14", "'Zone Area Matrix'!F14", 75, "Admin; breakout / meeting <25%"),
    "ENTERTAINMENT":     ("'Zone Area Matrix'!B15", "'Zone Area Matrix'!F15", 65, "High diversity required; single use max 65%"),
    "SKY ZONE":          ("'Zone Area Matrix'!B16", "'Zone Area Matrix'!F16", 60, "Observation + F&B; no single program >60%"),
}

for i, (zone, (gross_ref, net_ref, max_pct, note)) in enumerate(thresholds.items()):
    r = 5 + i
    fill = C_LIGHT if i % 2 == 0 else C_WHITE

    ws2.cell(r, 1, zone).font = Font(name="Arial", size=10, bold=True, color=C_GOLD)
    ws2.cell(r, 1).fill = hdr_fill(C_DARK)
    ws2.cell(r, 1).alignment = left()
    ws2.cell(r, 1).border = thin_border

    ws2.cell(r, 2, f"={gross_ref}").number_format = "#,##0"
    ws2.cell(r, 2).font = Font(name="Arial", size=10, color="0000CC")
    ws2.cell(r, 2).fill = PatternFill("solid", start_color=C_BLUE_IN, fgColor=C_BLUE_IN)
    ws2.cell(r, 2).alignment = center()
    ws2.cell(r, 2).border = thin_border

    ws2.cell(r, 3, f"={net_ref}").number_format = "#,##0"
    ws2.cell(r, 3).font = Font(name="Arial", size=10, color="006600")
    ws2.cell(r, 3).fill = PatternFill("solid", start_color=C_GREEN, fgColor=C_GREEN)
    ws2.cell(r, 3).alignment = center()
    ws2.cell(r, 3).border = thin_border

    ws2.cell(r, 4, max_pct / 100).number_format = "0%"
    ws2.cell(r, 4).font = Font(name="Arial", size=10, bold=True, color="0000CC")
    ws2.cell(r, 4).fill = PatternFill("solid", start_color=C_BLUE_IN, fgColor=C_BLUE_IN)
    ws2.cell(r, 4).alignment = center()
    ws2.cell(r, 4).border = thin_border

    ws2.cell(r, 5, note).font = Font(name="Arial", size=9, italic=True)
    ws2.cell(r, 5).fill = PatternFill("solid", start_color=fill, fgColor=fill)
    ws2.cell(r, 5).alignment = left()
    ws2.cell(r, 5).border = thin_border
    ws2.row_dimensions[r].height = 22

# JSON output cell
r_json = 5 + len(thresholds) + 2
ws2.merge_cells(f"A{r_json}:E{r_json}")
ws2[f"A{r_json}"] = "JSON FOR SPECKLE AUTOMATE  →  Copy the cell below into Input 4 (Program Threshold Matrix)"
ws2[f"A{r_json}"].font = h_font(10)
ws2[f"A{r_json}"].fill = hdr_fill(C_ACCENT)
ws2[f"A{r_json}"].alignment = center()
ws2.row_dimensions[r_json].height = 24

r_json2 = r_json + 1
ws2.merge_cells(f"A{r_json2}:E{r_json2}")
json_str = (
    '{"MEDICAL": 75, "HOTEL": 80, "TRANSPORTATION HQ": 70, '
    '"CORPORATE OFFICE": 75, "ADMIN OFFICE": 75, "ENTERTAINMENT": 65, "SKY ZONE": 60}'
)
ws2[f"A{r_json2}"] = json_str
ws2[f"A{r_json2}"].font = Font(name="Courier New", size=9, color="006600")
ws2[f"A{r_json2}"].fill = PatternFill("solid", start_color="EEF9EE", fgColor="EEF9EE")
ws2[f"A{r_json2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.row_dimensions[r_json2].height = 36

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – FLOOR AREA ESTIMATOR
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Floor Area Estimator")

ws3.merge_cells("A1:H1")
ws3["A1"] = "FLOOR AREA ESTIMATOR  —  Floors per Zone"
ws3["A1"].font = h_font(14)
ws3["A1"].fill = hdr_fill(C_DARK)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 34

ws3.merge_cells("A2:H2")
ws3["A2"] = "Enter number of floors per zone → area per floor and cumulative areas calculate automatically"
ws3["A2"].font = Font(name="Arial", size=9, italic=True, color=C_WHITE)
ws3["A2"].fill = hdr_fill(C_ACCENT)
ws3["A2"].alignment = center()
ws3.row_dimensions[2].height = 18

fe_headers = [
    "ZONE", "GROSS AREA\n(m²)", "NET USABLE\n(m²)", "NO. OF\nFLOORS",
    "GROSS / FLOOR\n(m²)", "NET / FLOOR\n(m²)",
    "CORE / FLOOR\n(m²)", "STRUCTURAL\n/ FLOOR (m²)"
]
fe_widths = [22, 16, 16, 12, 16, 16, 16, 18]
for ci, (h, w) in enumerate(zip(fe_headers, fe_widths), 1):
    ws3.cell(4, ci, h).font = h_font(10)
    ws3.cell(4, ci).fill = hdr_fill(C_MID)
    ws3.cell(4, ci).alignment = center()
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[4].height = 40

# zone gross refs from sheet 1
zone_gross_refs = [f"'Zone Area Matrix'!B{row_h+1+i}" for i in range(len(zones))]
zone_net_refs   = [f"'Zone Area Matrix'!F{row_h+1+i}" for i in range(len(zones))]
default_floors  = [5, 12, 8, 10, 8, 6, 3]   # sensible starting defaults

for i, (zone, _) in enumerate(zones):
    r = 5 + i
    fill = C_LIGHT if i % 2 == 0 else C_WHITE

    ws3.cell(r, 1, zone).font = Font(name="Arial", size=10, bold=True, color=C_GOLD)
    ws3.cell(r, 1).fill = hdr_fill(C_DARK)
    ws3.cell(r, 1).alignment = left()
    ws3.cell(r, 1).border = thin_border

    ws3.cell(r, 2, f"={zone_gross_refs[i]}").number_format = "#,##0"
    ws3.cell(r, 2).font = b_font()
    ws3.cell(r, 2).fill = PatternFill("solid", start_color=fill, fgColor=fill)
    ws3.cell(r, 2).alignment = center()
    ws3.cell(r, 2).border = thin_border

    ws3.cell(r, 3, f"={zone_net_refs[i]}").number_format = "#,##0"
    ws3.cell(r, 3).font = Font(name="Arial", size=10, color="006600")
    ws3.cell(r, 3).fill = PatternFill("solid", start_color=C_GREEN, fgColor=C_GREEN)
    ws3.cell(r, 3).alignment = center()
    ws3.cell(r, 3).border = thin_border

    # D – floors input (blue = editable)
    ws3.cell(r, 4, default_floors[i]).font = Font(name="Arial", size=10, bold=True, color="0000CC")
    ws3.cell(r, 4).fill = PatternFill("solid", start_color=C_BLUE_IN, fgColor=C_BLUE_IN)
    ws3.cell(r, 4).alignment = center()
    ws3.cell(r, 4).border = thin_border

    # E – gross per floor
    ws3.cell(r, 5, f"=B{r}/D{r}").number_format = "#,##0"
    ws3.cell(r, 5).font = b_font(bold=True)
    ws3.cell(r, 5).fill = PatternFill("solid", start_color=fill, fgColor=fill)
    ws3.cell(r, 5).alignment = center()
    ws3.cell(r, 5).border = thin_border

    # F – net per floor
    ws3.cell(r, 6, f"=C{r}/D{r}").number_format = "#,##0"
    ws3.cell(r, 6).font = Font(name="Arial", size=10, bold=True, color="006600")
    ws3.cell(r, 6).fill = PatternFill("solid", start_color=C_GREEN, fgColor=C_GREEN)
    ws3.cell(r, 6).alignment = center()
    ws3.cell(r, 6).border = thin_border

    # G – core per floor
    ws3.cell(r, 7, f"=B{r}*'Zone Area Matrix'!$B$5/D{r}").number_format = "#,##0"
    ws3.cell(r, 7).font = b_font()
    ws3.cell(r, 7).fill = PatternFill("solid", start_color=C_WARN, fgColor=C_WARN)
    ws3.cell(r, 7).alignment = center()
    ws3.cell(r, 7).border = thin_border

    # H – structural per floor
    ws3.cell(r, 8, f"=B{r}*'Zone Area Matrix'!$B$6/D{r}").number_format = "#,##0"
    ws3.cell(r, 8).font = b_font()
    ws3.cell(r, 8).fill = PatternFill("solid", start_color=C_WARN, fgColor=C_WARN)
    ws3.cell(r, 8).alignment = center()
    ws3.cell(r, 8).border = thin_border

    ws3.row_dimensions[r].height = 22

# totals
r_tot3 = 5 + len(zones)
ws3.cell(r_tot3, 1, "TOTAL / AVG").font = h_font(11)
ws3.cell(r_tot3, 1).fill = hdr_fill(C_GOLD)
ws3.cell(r_tot3, 1).alignment = center()
for ci, formula in [
    (2, f"=SUM(B5:B{r_tot3-1})"),
    (3, f"=SUM(C5:C{r_tot3-1})"),
    (4, f"=SUM(D5:D{r_tot3-1})"),
    (5, f"=AVERAGE(E5:E{r_tot3-1})"),
    (6, f"=AVERAGE(F5:F{r_tot3-1})"),
    (7, f"=AVERAGE(G5:G{r_tot3-1})"),
    (8, f"=AVERAGE(H5:H{r_tot3-1})"),
]:
    ws3.cell(r_tot3, ci, formula).font = h_font(11)
    ws3.cell(r_tot3, ci).fill = hdr_fill(C_DARK)
    ws3.cell(r_tot3, ci).number_format = "#,##0"
    ws3.cell(r_tot3, ci).alignment = center()
    ws3.cell(r_tot3, ci).border = thin_border
ws3.row_dimensions[r_tot3].height = 26

ws3.freeze_panes = "B5"

# ── save ─────────────────────────────────────────────────────────────────────
out = "/home/claude/ZoneAreaMatrix.xlsx"
wb.save(out)
print("saved:", out)
