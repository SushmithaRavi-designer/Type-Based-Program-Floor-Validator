"""
utils/csv_exporter.py
─────────────────────
Build and export the program validation results to Excel format.
"""

import csv
import io
import tempfile
from typing import List, Dict


COLUMNS = [
    "Level",
    "Program",
    "Area",
    "Status",
    "Area_OffPeak",
    "Area_Morning",
    "Area_Afternoon",
    "Area_Evening",
]


PROGRAM_HEX_COLORS = {
    "medical": "D5E8D4",       # light green
    "transit": "DAE8FC",       # light blue
    "retail": "FFE6CC",        # light orange
    "amenities": "E1D5E7",     # light purple
    "residential": "FFF2CC",   # light yellow
    "office": "F8CECC",        # light red
    "parking": "F5F5F5",       # light grey
    "voids": "E6E6E6",         # light grey
    "workadmin": "FDD9B5",     # pale peach
    "corporate": "B1DDF0",     # pale blue
}

def rows_to_csv(rows: List[Dict]) -> str:
    """
    Convert a list of row dicts to a UTF-8 CSV string.
    Uses QUOTE_MINIMAL to ensure proper escaping and Excel compatibility.
    Automatically detects all columns from the data.

    Parameters
    ----------
    rows : list of dicts - keys become column headers

    Returns
    -------
    str  Complete CSV text (header + data rows).
    """
    if not rows:
        return ""
    
    buf = io.StringIO()
    
    # Detect columns from first row, preserving order
    fieldnames = list(rows[0].keys()) if rows else COLUMNS
    
    # QUOTE_MINIMAL: quotes only when needed for special characters
    # dialect='excel': standard Windows/Mac/Linux Excel format
    writer = csv.DictWriter(
        buf, 
        fieldnames=fieldnames, 
        extrasaction="ignore",
        quoting=csv.QUOTE_MINIMAL,
        lineterminator='\r\n'  # Windows-style line endings for Excel
    )
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def rows_to_excel_multi_sheet(sheets_dict: dict, filename: str = None) -> str:
    """
    Convert multiple sheet dicts to a properly formatted Excel (.xlsx) file with multiple sheets.
    One sheet per occupancy group.
    
    Parameters
    ----------
    sheets_dict : dict with format {sheet_name: [rows_list]}
    filename : str, optional. If provided, writes to this file. Otherwise returns temp file path.
    
    Returns
    -------
    str  Path to the Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        # Remove default sheet
        if wb.sheetnames:
            wb.remove(wb.active)
        
        for sheet_name, rows in sheets_dict.items():
            # Sanitize sheet name (Excel has a 31-char limit)
            safe_sheet_name = str(sheet_name)[:31]
            ws = wb.create_sheet(title=safe_sheet_name)
            
            # Detect columns from first row, preserving order
            if rows and isinstance(rows[0], dict):
                fieldnames = list(rows[0].keys())
            else:
                fieldnames = COLUMNS
            
            # Write headers
            for col_idx, header in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                # Style header
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Write data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, header in enumerate(fieldnames, 1):
                    value = row_data.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    # Add background color for Program cells
                    if header == "Program" and value:
                        prog_key = str(value).lower().strip()
                        color_hex = PROGRAM_HEX_COLORS.get(prog_key)
                        if color_hex:
                            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                            
                    # Right align numeric columns
                    if "Area" in header or "Area_" in header:
                        try:
                            float(value)
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        except (ValueError, TypeError):
                            pass
                    # Center align Status column
                    if header == "Status":
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths dynamically
            for col_idx, header in enumerate(fieldnames, 1):
                if "Area" in header:
                    ws.column_dimensions[chr(64 + col_idx)].width = 18
                elif "Status" in header:
                    ws.column_dimensions[chr(64 + col_idx)].width = 35
                else:
                    ws.column_dimensions[chr(64 + col_idx)].width = 20
            
            # Set row height for header
            ws.row_dimensions[1].height = 25
        
        # Save to file
        if filename is None:
            # Create a temporary file
            tmp_file = tempfile.NamedTemporaryFile(
                suffix=".xlsx",
                delete=False,
                prefix="program_floor_validation_"
            )
            filename = tmp_file.name
            tmp_file.close()
        
        wb.save(filename)
        return filename
    
    except ImportError as ex:
        raise RuntimeError(
            "Multi-sheet Excel export requires 'openpyxl'. "
            "Install dependencies from pyproject.toml/requirements.txt before running."
        ) from ex


def rows_to_excel(rows: List[Dict], filename: str = None) -> str:
    """
    Convert a list of row dicts to a properly formatted Excel (.xlsx) file.
    Falls back to CSV if openpyxl is not available.
    Automatically detects columns from the data.
    
    Parameters
    ----------
    rows : list of dicts - keys become column headers
    filename : str, optional. If provided, writes to this file. Otherwise returns temp file path.
    
    Returns
    -------
    str  Path to the Excel/CSV file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Program Floor Analysis"
        
        # Detect columns from first row, preserving order
        fieldnames = list(rows[0].keys()) if rows else COLUMNS
        
        # Write headers
        for col_idx, header in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # Style header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(fieldnames, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Right align numeric columns
                if "Area" in header or "Area_" in header:
                    try:
                        float(value)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    except (ValueError, TypeError):
                        pass
                # Center align Status column
                if header == "Status":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths dynamically
        for col_idx, header in enumerate(fieldnames, 1):
            if "Area" in header:
                ws.column_dimensions[chr(64 + col_idx)].width = 18
            elif "Status" in header:
                ws.column_dimensions[chr(64 + col_idx)].width = 35
            else:
                ws.column_dimensions[chr(64 + col_idx)].width = 20
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        # Save to file
        if filename is None:
            # Create a temporary file
            tmp_file = tempfile.NamedTemporaryFile(
                suffix=".xlsx",
                delete=False,
                prefix="program_floor_validation_"
            )
            filename = tmp_file.name
            tmp_file.close()
        
        wb.save(filename)
        return filename
    
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        buf = io.StringIO()
        
        # Detect columns from first row
        fieldnames = list(rows[0].keys()) if rows else COLUMNS
        
        writer = csv.DictWriter(
            buf, 
            fieldnames=fieldnames, 
            extrasaction="ignore",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator='\r\n'
        )
        writer.writeheader()
        writer.writerows(rows)
        
        if filename is None:
            tmp_file = tempfile.NamedTemporaryFile(
                suffix=".csv",
                delete=False,
                prefix="program_floor_validation_"
            )
            filename = tmp_file.name
            tmp_file.close()
        
        with open(filename, 'w') as f:
            f.write(buf.getvalue())
        
        return filename


def rows_to_floor_summary_csv(floor_data: dict, stacking: dict, thresholds: dict,
                               default_threshold: float) -> str:
    """
    Export a compact per-floor summary CSV (one row per floor).

    Columns: Floor | Total Area | Programs | Dominant | Dominant % | Diversity H | Stacking
    """
    from utils.kpi import floor_summary

    buf = io.StringIO()
    summary_cols = [
        "Floor", "Total Area (m²)", "# Programs", "Dominant Program",
        "Dominant %", "Allowed %", "Mono-Functional?", "Diversity Index (H)",
    ]
    writer = csv.DictWriter(buf, fieldnames=summary_cols)
    writer.writeheader()

    import re
    def _level_sort_key(name):
        m = re.search(r'(\d+)', str(name))
        return (int(m.group(1)) if m else 0, str(name))

    for floor, prog_areas in sorted(floor_data.items(), key=lambda x: _level_sort_key(x[0])):
        s = floor_summary(floor, prog_areas, thresholds, default_threshold)
        writer.writerow({
            "Floor":               s["floor"],
            "Total Area (m²)":     s["total_area"],
            "# Programs":          s["num_programs"],
            "Dominant Program":    s["dominant_program"],
            "Dominant %":          s["dominant_pct"],
            "Allowed %":           s["allowed_pct"],
            "Mono-Functional?":    "YES" if s["is_mono_functional"] else "NO",
            "Diversity Index (H)": s["diversity_index"],
        })
    return buf.getvalue()
