from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── palette (match existing workbook) ────────────────────────────────────────
C_DARK   = "1A1A2E"
C_MID    = "16213E"
C_ACCENT = "0F3460"
C_GOLD   = "E94560"
C_LIGHT  = "F5F5F5"
C_WHITE  = "FFFFFF"
C_BLUE   = "D6E4F0"
C_WARN   = "FFF3CD"
C_GREEN  = "D4EDDA"
C_RED    = "F8D7DA"

def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def s(st="thin", c="BBBBBB"): return Side(style=st, color=c)
thin  = Border(left=s(), right=s(), top=s(), bottom=s())
thick = Border(left=s(), right=s(), top=s("medium","333333"), bottom=s("medium","333333"))
def hf(sz=10, bold=True, col=C_WHITE): return Font(name="Arial", size=sz, bold=bold, color=col)
def bf(sz=10, bold=False, col="222222"): return Font(name="Arial", size=sz, bold=bold, color=col)
def c_al(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def l_al(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── zone definitions ─────────────────────────────────────────────────────────
zones = [
    ("MEDICAL",           35000, ["B1","B2","B3","B4","B5"]),
    ("HOTEL",             50000, ["L1","L2","L3","L4","L5","L6","L7","L8","L9","L10","L11","L12"]),
    ("TRANSPORTATION HQ", 52000, ["T1","T2","T3","T4","T5","T6","T7","T8"]),
    ("CORPORATE OFFICE",  30000, ["C1","C2","C3","C4","C5","C6","C7","C8","C9","C10"]),
    ("ADMIN OFFICE",      30000, ["A1","A2","A3","A4","A5","A6","A7","A8"]),
    ("ENTERTAINMENT",     31000, ["E1","E2","E3","E4","E5","E6"]),
    ("SKY ZONE",          10400, ["S1","S2","S3"]),
]

wb = load_workbook("/home/claude/ZoneAreaMatrix.xlsx")

# ════════════════════════════════════════════════════════════════════════════
# NEW SHEET – FLOOR-BY-FLOOR BREAKDOWN
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Floor Breakdown", 2)   # insert as 3rd sheet

# ── title ────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:J1")
ws["A1"] = "FLOOR-BY-FLOOR AREA BREAKDOWN  —  Core 15% + Structural 15% per Floor"
ws["A1"].font = hf(14)
ws["A1"].fill = fill(C_DARK)
ws["A1"].alignment = c_al()
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:J2")
ws["A2"] = (
    "Each floor's gross area = Zone Gross ÷ No. of Floors  |  "
    "Core = 15%  |  Structural = 15%  |  Net Usable = 70%"
)
ws["A2"].font = Font(name="Arial", size=9, italic=True, color=C_WHITE)
ws["A2"].fill = fill(C_ACCENT)
ws["A2"].alignment = c_al()
ws.row_dimensions[2].height = 18

# ── global assumption cells (row 4) ─────────────────────────────────────────
ws["A4"] = "Core %"
ws["B4"] = 0.15
ws["A4"].font = hf(10, True, "222222")
ws["A4"].fill = fill(C_WARN)
ws["A4"].alignment = l_al()
ws["B4"].font = Font(name="Arial", size=10, bold=True, color="0000CC")
ws["B4"].fill = fill(C_BLUE)
ws["B4"].number_format = "0%"
ws["B4"].alignment = c_al()

ws["C4"] = "Structural %"
ws["D4"] = 0.15
ws["C4"].font = hf(10, True, "222222")
ws["C4"].fill = fill(C_WARN)
ws["C4"].alignment = l_al()
ws["D4"].font = Font(name="Arial", size=10, bold=True, color="0000CC")
ws["D4"].fill = fill(C_BLUE)
ws["D4"].number_format = "0%"
ws["D4"].alignment = c_al()

ws["E4"] = "Net Usable %"
ws["F4"] = "=1-B4-D4"
ws["E4"].font = hf(10, True, "222222")
ws["E4"].fill = fill(C_WARN)
ws["E4"].alignment = l_al()
ws["F4"].font = Font(name="Arial", size=10, bold=True, color="006600")
ws["F4"].fill = fill(C_GREEN)
ws["F4"].number_format = "0%"
ws["F4"].alignment = c_al()

ws.row_dimensions[4].height = 22

# ── column widths ────────────────────────────────────────────────────────────
col_w = [22, 6, 16, 16, 16, 16, 16, 16, 16, 16]
col_names = [
    "ZONE / FLOOR", "FL\n#",
    "GROSS AREA\n/ FLOOR (m²)",
    "CORE AREA\n(15%) (m²)",
    "STRUCTURAL\n(15%) (m²)",
    "TOTAL\nDEDUCTION (m²)",
    "NET USABLE\n(70%) (m²)",
    "CORE %", "STRUCT %", "NET %"
]
for ci, (h, w) in enumerate(zip(col_names, col_w), 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# ── header row ───────────────────────────────────────────────────────────────
HDR = 6
for ci, h in enumerate(col_names, 1):
    c = ws.cell(HDR, ci, h)
    c.font = hf(9)
    c.fill = fill(C_MID)
    c.alignment = c_al()
    c.border = thin
ws.row_dimensions[HDR].height = 44

# ── data: one zone block per zone, one row per floor ─────────────────────────
cur_row = HDR + 1
zone_summary_refs = []   # (zone_name, gross, first_floor_row, last_floor_row)

for z_idx, (zone_name, gross_total, floor_labels) in enumerate(zones):
    n_floors = len(floor_labels)
    first_r  = cur_row
    gross_per_floor_formula_base = gross_total  # hardcode gross; formulas derive everything

    # ── zone label (merged across all floor rows) ────────────────────────────
    if n_floors > 1:
        ws.merge_cells(f"A{first_r}:A{first_r + n_floors - 1}")

    zone_cell = ws.cell(first_r, 1, zone_name)
    zone_cell.font  = Font(name="Arial", size=10, bold=True, color=C_GOLD)
    zone_cell.fill  = fill(C_DARK)
    zone_cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True, text_rotation=0)
    zone_cell.border = thick

    # ── floor rows ────────────────────────────────────────────────────────────
    for f_idx, floor_label in enumerate(floor_labels):
        r       = cur_row + f_idx
        row_alt = C_LIGHT if f_idx % 2 == 0 else C_WHITE

        # B – floor number
        c = ws.cell(r, 2, floor_label)
        c.font = Font(name="Arial", size=8, bold=True, color=C_WHITE)
        c.fill = fill(C_ACCENT)
        c.alignment = c_al()
        c.border = thin

        # C – gross per floor  = zone_total / n_floors  (formula)
        gross_ref = gross_total   # we store the gross total as a number for formula reference
        ws.cell(r, 3, f"={gross_total}/{n_floors}").font = Font(name="Arial",size=10,color="0000CC")
        ws.cell(r, 3).fill = fill(C_BLUE)
        ws.cell(r, 3).number_format = "#,##0.0"
        ws.cell(r, 3).alignment = c_al()
        ws.cell(r, 3).border = thin

        # D – Core 15%
        ws.cell(r, 4, f"=C{r}*$B$4").font = bf()
        ws.cell(r, 4).fill = fill(C_WARN)
        ws.cell(r, 4).number_format = "#,##0.0"
        ws.cell(r, 4).alignment = c_al()
        ws.cell(r, 4).border = thin

        # E – Structural 15%
        ws.cell(r, 5, f"=C{r}*$D$4").font = bf()
        ws.cell(r, 5).fill = fill(C_WARN)
        ws.cell(r, 5).number_format = "#,##0.0"
        ws.cell(r, 5).alignment = c_al()
        ws.cell(r, 5).border = thin

        # F – Total deduction
        ws.cell(r, 6, f"=D{r}+E{r}").font = Font(name="Arial",size=10,bold=True,color="CC0000")
        ws.cell(r, 6).fill = fill("FFE0E0")
        ws.cell(r, 6).number_format = "#,##0.0"
        ws.cell(r, 6).alignment = c_al()
        ws.cell(r, 6).border = thin

        # G – Net usable 70%
        ws.cell(r, 7, f"=C{r}*$F$4").font = Font(name="Arial",size=10,bold=True,color="006600")
        ws.cell(r, 7).fill = fill(C_GREEN)
        ws.cell(r, 7).number_format = "#,##0.0"
        ws.cell(r, 7).alignment = c_al()
        ws.cell(r, 7).border = thin

        # H/I/J – percentage labels (formula-locked to assumptions)
        for ci, ref in [(8,"$B$4"),(9,"$D$4"),(10,"$F$4")]:
            ws.cell(r, ci, f"={ref}").font = bf(9)
            ws.cell(r, ci).fill = fill(row_alt)
            ws.cell(r, ci).number_format = "0%"
            ws.cell(r, ci).alignment = c_al()
            ws.cell(r, ci).border = thin

        ws.row_dimensions[r].height = 20

    # ── zone subtotal row ─────────────────────────────────────────────────────
    sub_r = cur_row + n_floors
    ws.cell(sub_r, 1, f"↳ {zone_name} TOTAL").font = hf(9, True, C_GOLD)
    ws.cell(sub_r, 1).fill = fill(C_MID)
    ws.cell(sub_r, 1).alignment = l_al()
    ws.cell(sub_r, 1).border = thick

    for ci, (fmt, is_sum) in enumerate([
        ("#,##0.0", True),   # C
        ("#,##0.0", True),   # D
        ("#,##0.0", True),   # E
        ("#,##0.0", True),   # F
        ("#,##0.0", True),   # G
        ("0%", False),       # H
        ("0%", False),       # I
        ("0%", False),       # J
    ], 3):
        col_l = get_column_letter(ci)
        formula = (f"=SUM({col_l}{first_r}:{col_l}{sub_r-1})"
                   if is_sum else "")
        ref_map = {8: "=$B$4", 9: "=$D$4", 10: "=$F$4"}
        if ci in ref_map:
            formula = ref_map[ci]

        ws.cell(sub_r, ci, formula if formula else "").font = hf(9, True, C_WHITE)
        ws.cell(sub_r, ci).fill = fill(C_MID)
        ws.cell(sub_r, ci).number_format = fmt
        ws.cell(sub_r, ci).alignment = c_al()
        ws.cell(sub_r, ci).border = thick

    ws.cell(sub_r, 2, f"{n_floors} fl.").font = hf(8, False, C_WHITE)
    ws.cell(sub_r, 2).fill = fill(C_MID)
    ws.cell(sub_r, 2).alignment = c_al()
    ws.cell(sub_r, 2).border = thick
    ws.row_dimensions[sub_r].height = 22

    zone_summary_refs.append((zone_name, gross_total, n_floors, first_r, sub_r))
    cur_row = sub_r + 2   # blank gap between zones

# ── GRAND TOTAL row ───────────────────────────────────────────────────────────
gt_r = cur_row + 1
ws.merge_cells(f"A{gt_r}:B{gt_r}")
ws.cell(gt_r, 1, "GRAND TOTAL — ALL ZONES").font = hf(11, True, C_WHITE)
ws.cell(gt_r, 1).fill = fill(C_GOLD)
ws.cell(gt_r, 1).alignment = c_al()
ws.cell(gt_r, 1).border = thick

# Sum only the subtotal rows for each column
sub_refs_C = "+".join([f"C{ref[4]}" for ref in zone_summary_refs])
sub_refs_D = "+".join([f"D{ref[4]}" for ref in zone_summary_refs])
sub_refs_E = "+".join([f"E{ref[4]}" for ref in zone_summary_refs])
sub_refs_F = "+".join([f"F{ref[4]}" for ref in zone_summary_refs])
sub_refs_G = "+".join([f"G{ref[4]}" for ref in zone_summary_refs])

for ci, formula in [
    (3, f"={sub_refs_C}"),
    (4, f"={sub_refs_D}"),
    (5, f"={sub_refs_E}"),
    (6, f"={sub_refs_F}"),
    (7, f"={sub_refs_G}"),
]:
    ws.cell(gt_r, ci, formula).font = hf(11)
    ws.cell(gt_r, ci).fill = fill(C_DARK)
    ws.cell(gt_r, ci).number_format = "#,##0.0"
    ws.cell(gt_r, ci).alignment = c_al()
    ws.cell(gt_r, ci).border = thick

for ci in (8, 9, 10):
    ref = {8:"=$B$4", 9:"=$D$4", 10:"=$F$4"}[ci]
    ws.cell(gt_r, ci, ref).font = hf(11)
    ws.cell(gt_r, ci).fill = fill(C_DARK)
    ws.cell(gt_r, ci).number_format = "0%"
    ws.cell(gt_r, ci).alignment = c_al()
    ws.cell(gt_r, ci).border = thick

ws.row_dimensions[gt_r].height = 28
ws.freeze_panes = "C7"

out = "/home/claude/ZoneAreaMatrix.xlsx"
wb.save(out)
print("saved:", out)
